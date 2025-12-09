# backend/excel_exporter.py

import pandas as pd
import io
import logging
from openpyxl.styles import PatternFill, Font

# Настройка логгера
log = logging.getLogger(__name__)


def _auto_adjust_column_width(worksheet):
    """Автоподбор ширины столбцов на листе Excel."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Пытаемся получить длину заголовка, если он есть
        try:
            header_length = len(str(column_cells[0].value)) if column_cells[0].value else 0
        except:
            header_length = 0

        adjusted_width = max(max_length + 2, header_length + 2)
        # Ограничим максимальную ширину, чтобы не было гигантских колонок
        if adjusted_width > 100:
            adjusted_width = 100
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _color_header_by_name(worksheet, dataframe, column_name, fill_style):
    """Раскрашивает заголовок столбца по его имени."""
    try:
        column_names = dataframe.columns.tolist()
        if column_name in column_names:
            col_index = column_names.index(column_name) + 1
            header_cell = worksheet.cell(row=1, column=col_index)
            header_cell.fill = fill_style
    except Exception as e:
        log.warning("Не удалось раскрасить заголовок '%s': %s", column_name, e)


def _add_autofilter(worksheet):
    """Добавляет автофильтр на лист Excel."""
    if worksheet.max_row > 1:
        try:
            worksheet.auto_filter.ref = worksheet.dimensions
        except Exception as e:
            log.warning("Не удалось добавить автофильтр: %s", e)


def export_results_to_stream(results_to_export):
    """
    Экспортирует словарь с результатами (DataFrame) в поток байтов (в память).
    Возвращает объект io.BytesIO.
    """

    if not results_to_export:
        log.error("Вызов export_results с пустым словарем results_to_export.")
        raise ValueError("Нет данных для экспорта.")

    # Создаем поток в памяти вместо файла на диске
    output_stream = io.BytesIO()

    log.info("Начало генерации Excel в памяти...")

    # Используем writer с потоком
    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        # --- Стили ---
        blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        headers_to_color = ["Account", "Субсчет в учетной организации"]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)

        def write_sheet(sheet_name, df_key):
            """Вспомогательная функция для записи одного листа."""
            log.debug("Запись листа: '%s' (ключ: %s)", sheet_name, df_key)
            df = results_to_export.get(df_key)

            # Если данных нет или None, создаем пустой DataFrame
            if df is None:
                df = pd.DataFrame()

            # Если пришел список словарей (JSON), конвертируем в DataFrame
            if isinstance(df, list):
                df = pd.DataFrame(df)

            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Получаем доступ к листу для форматирования
            worksheet = writer.sheets[sheet_name]
            _auto_adjust_column_width(worksheet)
            _add_autofilter(worksheet)

            for header in headers_to_color:
                _color_header_by_name(worksheet, df, header, blue_fill)

        # --- 1. Основные листы ---
        write_sheet('Совпадения', 'matches')
        write_sheet('Расхождения_Unity', 'unmatched1')
        write_sheet('Расхождения_АИС', 'unmatched2')

        # --- 2. Лист "Сводка" (с выделением) ---
        log.debug("Форматирование листа 'Сводка'...")
        summary1 = results_to_export.get('summary1')
        summary2 = results_to_export.get('summary2')

        # Конвертация из JSON (dict) в DataFrame, если нужно
        if isinstance(summary1, dict): summary1 = pd.Series(summary1)
        if isinstance(summary2, dict): summary2 = pd.Series(summary2)

        if summary1 is not None and summary2 is not None and (not summary1.empty or not summary2.empty):
            summary_df = pd.concat([summary1, summary2], axis=1)
            summary_df.columns = ['Unity', 'АИС']
            summary_df = summary_df.fillna(0).astype(int)

            summary_df.to_excel(writer, sheet_name='Сводка', index=True)
            worksheet_summary = writer.sheets['Сводка']
            _auto_adjust_column_width(worksheet_summary)

            # Подсветка расхождений
            for index, row_data in summary_df.iterrows():
                if row_data['Unity'] != row_data['АИС']:
                    # +2 потому что заголовки занимают строку и индексация с 1
                    try:
                        excel_row_num = summary_df.index.get_loc(index) + 2
                        for col_num in range(1, worksheet_summary.max_column + 1):
                            worksheet_summary.cell(row=excel_row_num, column=col_num).fill = yellow_fill
                    except Exception as e:
                        pass  # Игнорируем ошибки индексации при форматировании

        # --- 3. Листы "Задвоения" ---
        log.debug("Проверка и запись листов 'Задвоения'...")
        duplicates_df1 = results_to_export.get('duplicates1')
        if duplicates_df1 is not None:
            if isinstance(duplicates_df1, list): duplicates_df1 = pd.DataFrame(duplicates_df1)
            if not duplicates_df1.empty:
                write_sheet('Задвоения_Unity', 'duplicates1')

        duplicates_df2 = results_to_export.get('duplicates2')
        if duplicates_df2 is not None:
            if isinstance(duplicates_df2, list): duplicates_df2 = pd.DataFrame(duplicates_df2)
            if not duplicates_df2.empty:
                write_sheet('Задвоения_АИС', 'duplicates2')

        # --- 4. Специальный лист "ПОДФТ" (комбинированный) ---
        log.debug("Форматирование специального листа 'ПОДФТ'...")

        raw_podft_7m = results_to_export.get('podft_7m_deals', [])
        podft_7m_df = pd.DataFrame(raw_podft_7m) if isinstance(raw_podft_7m,
                                                               list) else raw_podft_7m.copy() if raw_podft_7m is not None else pd.DataFrame()

        podft_7m_df.to_excel(writer, sheet_name='ПОДФТ', index=False)
        worksheet_podft = writer.sheets['ПОДФТ']

        _auto_adjust_column_width(worksheet_podft)
        _add_autofilter(worksheet_podft)
        for header in headers_to_color:
            _color_header_by_name(worksheet_podft, podft_7m_df, header, blue_fill)

        date_counts_7m = None
        if not podft_7m_df.empty and 'Дата валютирования' in podft_7m_df.columns:
            try:
                podft_7m_df['date_only'] = pd.to_datetime(podft_7m_df['Дата валютирования']).dt.date
                date_counts_7m = podft_7m_df['date_only'].value_counts()
                if not date_counts_7m.empty and len(date_counts_7m) > 1:
                    majority_date = date_counts_7m.idxmax()
                    date_col_idx = podft_7m_df.columns.get_loc('Дата валютирования') + 1
                    for row in range(2, worksheet_podft.max_row + 1):
                        cell_value = worksheet_podft.cell(row=row, column=date_col_idx).value
                        if cell_value:
                            cell_date = pd.to_datetime(cell_value).date()
                            if cell_date != majority_date:
                                for col in range(1, podft_7m_df.shape[1] + 1):
                                    worksheet_podft.cell(row=row, column=col).fill = yellow_fill
            except Exception as e:
                log.warning("Ошибка при выделении дат в ПОДФТ (7М): %s", e)

        # Статистика 7М
        summary_row_index = worksheet_podft.max_row + 2
        total_cell = worksheet_podft.cell(row=summary_row_index, column=1,
                                          value=f"Общее количество (>= 7M): {len(podft_7m_df)}")
        total_cell.font = bold_font

        if date_counts_7m is not None and not date_counts_7m.empty:
            group_header_cell = worksheet_podft.cell(row=summary_row_index + 2, column=1,
                                                     value="Количество по датам (>= 7M):")
            group_header_cell.font = bold_font
            current_row = summary_row_index + 3
            for date_val, count in date_counts_7m.items():
                worksheet_podft.cell(row=current_row, column=1, value=date_val.strftime('%Y-%m-%d'))
                worksheet_podft.cell(row=current_row, column=2, value=count)
                current_row += 1

        # --- 4.1 Часть 45M (Бонды/Опционы) на том же листе ---
        raw_podft_45m = results_to_export.get('podft_45m_bo_deals', [])
        podft_45m_bo_df = pd.DataFrame(raw_podft_45m) if isinstance(raw_podft_45m,
                                                                    list) else raw_podft_45m.copy() if raw_podft_45m is not None else pd.DataFrame()

        separator_row = worksheet_podft.max_row + 3
        sep_cell = worksheet_podft.cell(row=separator_row, column=1,
                                        value="--- ПОДФТ: БОНДЫ И ОПЦИОНЫ (>= 45 000 000) ---")
        sep_cell.font = bold_font
        table_2_start_row = separator_row

        if not podft_45m_bo_df.empty:
            podft_45m_bo_df.to_excel(writer, sheet_name='ПОДФТ', index=False, startrow=table_2_start_row + 1)
            header_row_2 = table_2_start_row + 2

            # Раскраска заголовков второй таблицы
            for header in headers_to_color:
                try:
                    column_names = podft_45m_bo_df.columns.tolist()
                    if header in column_names:
                        col_index = column_names.index(header) + 1
                        header_cell = worksheet_podft.cell(row=header_row_2, column=col_index)
                        header_cell.fill = blue_fill
                except Exception as e:
                    pass

            date_counts_bo = None
            if 'Дата валютирования' in podft_45m_bo_df.columns:
                try:
                    podft_45m_bo_df['date_only'] = pd.to_datetime(podft_45m_bo_df['Дата валютирования']).dt.date
                    date_counts_bo = podft_45m_bo_df['date_only'].value_counts()
                    if not date_counts_bo.empty and len(date_counts_bo) > 1:
                        majority_date_bo = date_counts_bo.idxmax()
                        date_col_idx_bo = podft_45m_bo_df.columns.get_loc('Дата валютирования') + 1
                        for row in range(header_row_2 + 1, worksheet_podft.max_row + 1):
                            cell_value = worksheet_podft.cell(row=row, column=date_col_idx_bo).value
                            if cell_value:
                                cell_date = pd.to_datetime(cell_value).date()
                                if cell_date != majority_date_bo:
                                    for col in range(1, podft_45m_bo_df.shape[1] + 1):
                                        worksheet_podft.cell(row=row, column=col).fill = yellow_fill
                except Exception as e:
                    log.warning("Ошибка при выделении дат в ПОДФТ (Бонды/Опционы): %s", e)

            # Статистика 45М
            summary_row_index_2 = worksheet_podft.max_row + 2
            total_cell_2 = worksheet_podft.cell(row=summary_row_index_2, column=1,
                                                value=f"Общее количество (Бонды/Опционы): {len(podft_45m_bo_df)}")
            total_cell_2.font = bold_font

            if date_counts_bo is not None and not date_counts_bo.empty:
                group_header_cell_2 = worksheet_podft.cell(row=summary_row_index_2 + 2, column=1,
                                                           value="Количество по датам (Бонды/Опционы):")
                group_header_cell_2.font = bold_font
                current_row_2 = summary_row_index_2 + 3
                for date_val, count in date_counts_bo.items():
                    worksheet_podft.cell(row=current_row_2, column=1, value=date_val.strftime('%Y-%m-%d'))
                    worksheet_podft.cell(row=current_row_2, column=2, value=count)
                    current_row_2 += 1
        else:
            worksheet_podft.cell(row=separator_row + 2, column=1,
                                 value="Сделок по Бондам и Опционам (>= 45М) не найдено.")

        # --- 5. Специальный лист "КРИПТО" (с фильтрацией >= 5М) ---
        log.debug("Форматирование специального листа 'КРИПТО'...")

        raw_crypto = results_to_export.get('crypto_deals')
        crypto_deals_df = pd.DataFrame(raw_crypto) if isinstance(raw_crypto,
                                                                 list) else raw_crypto.copy() if raw_crypto is not None else pd.DataFrame()

        if not crypto_deals_df.empty and 'Сумма тг' in crypto_deals_df.columns:
            # Преобразуем в числа, если это еще не числа (актуально после JSON)
            crypto_deals_df['Сумма тг'] = pd.to_numeric(crypto_deals_df['Сумма тг'], errors='coerce')
            high_value_crypto_df = crypto_deals_df[crypto_deals_df['Сумма тг'] >= 5000000].copy()
        else:
            high_value_crypto_df = pd.DataFrame(columns=crypto_deals_df.columns if not crypto_deals_df.empty else [])

        high_value_crypto_df.to_excel(writer, sheet_name='КРИПТО', index=False)
        worksheet_crypto = writer.sheets['КРИПТО']

        _auto_adjust_column_width(worksheet_crypto)
        _add_autofilter(worksheet_crypto)
        for header in headers_to_color:
            _color_header_by_name(worksheet_crypto, high_value_crypto_df, header, blue_fill)

        date_counts_crypto = None
        if not high_value_crypto_df.empty and 'Дата валютирования' in high_value_crypto_df.columns:
            try:
                high_value_crypto_df['date_only'] = pd.to_datetime(
                    high_value_crypto_df['Дата валютирования']).dt.date
                date_counts_crypto = high_value_crypto_df['date_only'].value_counts()
                if not date_counts_crypto.empty and len(date_counts_crypto) > 1:
                    majority_date_crypto = date_counts_crypto.idxmax()
                    date_col_idx_crypto = high_value_crypto_df.columns.get_loc('Дата валютирования') + 1
                    for row in range(2, worksheet_crypto.max_row + 1):
                        cell_value = worksheet_crypto.cell(row=row, column=date_col_idx_crypto).value
                        if cell_value:
                            cell_date = pd.to_datetime(cell_value).date()
                            if cell_date != majority_date_crypto:
                                for col in range(1, high_value_crypto_df.shape[1] + 1):
                                    worksheet_crypto.cell(row=row, column=col).fill = yellow_fill
            except Exception as e:
                log.warning("Ошибка при выделении дат в КРИПТО: %s", e)

        summary_row_index = worksheet_crypto.max_row + 2
        count_cell = worksheet_crypto.cell(row=summary_row_index, column=1,
                                           value=f"Количество: {len(high_value_crypto_df)}")
        count_cell.font = bold_font
        text_cell = worksheet_crypto.cell(row=summary_row_index, column=2,
                                          value="КРИПТО СДЕЛКИ >= 5 000 000 тг")
        text_cell.font = bold_font

        if date_counts_crypto is not None and not date_counts_crypto.empty:
            group_header_cell = worksheet_crypto.cell(row=summary_row_index + 2, column=1,
                                                      value="Количество по датам:")
            group_header_cell.font = bold_font
            current_row = summary_row_index + 3
            for date_val, count in date_counts_crypto.items():
                worksheet_crypto.cell(row=current_row, column=1, value=date_val.strftime('%Y-%m-%d'))
                worksheet_crypto.cell(row=current_row, column=2, value=count)
                current_row += 1

    log.info("Экспорт в память успешно завершен.")

    # Сбрасываем курсор потока в начало, чтобы файл можно было прочитать
    output_stream.seek(0)
    return output_stream