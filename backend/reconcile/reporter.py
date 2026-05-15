from __future__ import annotations

import hashlib
import re
from typing import Any, Dict, List, Optional, Set

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import ReconcileParams, ReconcileSummary

SHEET_SUMMARY = "Сводка"
SHEET_MATCHES = "Совпадения"
SHEET_MISSING = "Нет в Unity"
SHEET_EXTRA = "Лишнее в Unity"
SHEET_UNITY_STATUS = "Статус Unity"
SHEET_VOL_SYMBOL = "Объем Инстр"
SHEET_VOL_SYMBOL_SIDE = "Объем Инстр+Side"


def _cols(df: pd.DataFrame, ordered: List[str]) -> List[str]:
    return [c for c in ordered if c in df.columns]


def _safe_rename(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    m = {k: v for k, v in mapping.items() if k in df.columns}
    return df.rename(columns=m)


def _make_unique_headers(cols: List[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in cols:
        base = str(c).strip() if c is not None else ""
        if not base:
            base = "col"
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
        out.append(base)
    return out


def _clean_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = _make_unique_headers(list(out.columns))
    out = out.replace([np.inf, -np.inf], np.nan)
    for c in out.columns:
        if out[c].dtype == object:
            out[c] = out[c].replace({"nan": "", "NaN": "", "NaT": "", "None": ""})
    return out


def _make_table_name(ws_title: str, idx: int) -> str:
    base = re.sub(r"[^A-Za-z0-9_]", "_", ws_title).strip("_")
    if not base:
        base = "Sheet"
    base = base[:25]
    h = hashlib.md5(ws_title.encode("utf-8")).hexdigest()[:6]
    name = f"T_{base}_{idx}_{h}"
    return name[:255]


def _build_pretty_tables(
    matched_all: pd.DataFrame,
    exchange_n: pd.DataFrame,
    unity_n: pd.DataFrame,
    missing_in_unity: pd.DataFrame,
    extra_in_unity: pd.DataFrame,
    ex_status: pd.DataFrame,
    uni_status: pd.DataFrame,
    volume_by_symbol: Optional[pd.DataFrame],
    volume_by_symbol_side: Optional[pd.DataFrame],
    *,
    exchange_name: str,
    params: ReconcileParams,
) -> Dict[str, Optional[pd.DataFrame]]:
    exn = exchange_name.strip().lower()
    prefix = "B" if exn == "binance" else ("O" if exn == "okx" else ("Y" if exn == "bybit" else "X"))
    p = f"{prefix}_"

    ex_sel = exchange_n.reset_index().rename(columns={"index": "exchange_idx"})
    u_sel = unity_n.reset_index().rename(columns={"index": "unity_idx"})

    ex_keep = _cols(ex_sel, [
        "exchange_idx",
        "Trade ID", "Order ID", "Insert Time",
        "trade_dt_utc",
        "symbol", "side", "qty", "price", "notional",
        "Fee", "Commission Asset",
    ])
    u_keep = _cols(u_sel, [
        "unity_idx",
        "ID", "Transact time",
        "trade_dt_utc",
        "Instrument",
        "symbol", "side", "qty", "price", "notional",
        "Net commission amount",
    ])

    ex_view = ex_sel[ex_keep].copy()
    u_view = u_sel[u_keep].copy()

    ex_view = _safe_rename(ex_view, {
        "Trade ID": f"{p}TradeID",
        "Order ID": f"{p}OrderID",
        "Insert Time": f"{p}Время",
        "trade_dt_utc": f"{p}UTC",
        "symbol": f"{p}Символ",
        "side": f"{p}Сторона",
        "qty": f"{p}Qty",
        "price": f"{p}Цена",
        "notional": f"{p}Объем",
        "Fee": f"{p}Fee",
        "Commission Asset": f"{p}Комиссия_Asset",
    })

    u_view = _safe_rename(u_view, {
        "ID": "U_ID",
        "Transact time": "U_Время",
        "trade_dt_utc": "U_UTC",
        "Instrument": "U_Instrument",
        "symbol": "U_Символ",
        "side": "U_Сторона",
        "qty": "U_Qty",
        "price": "U_Цена",
        "notional": "U_Объем",
        "Net commission amount": "U_Комиссия",
    })

    m = matched_all.copy()
    if "score" not in m.columns:
        m["score"] = np.nan
    if "key_used" not in m.columns:
        m["key_used"] = ""

    m = m.merge(ex_view, on="exchange_idx", how="left").merge(u_view, on="unity_idx", how="left")

    ex_utc = f"{p}UTC"
    ex_qty = f"{p}Qty"
    ex_price = f"{p}Цена"
    ex_not = f"{p}Объем"

    if ex_utc in m.columns and "U_UTC" in m.columns:
        t1 = pd.to_datetime(m[ex_utc], errors="coerce")
        t2 = pd.to_datetime(m["U_UTC"], errors="coerce")
        m["Δt_sec"] = (t1 - t2).abs().dt.total_seconds()

    if ex_qty in m.columns and "U_Qty" in m.columns:
        m["ΔQty"] = pd.to_numeric(m[ex_qty], errors="coerce") - pd.to_numeric(m["U_Qty"], errors="coerce")

    if ex_price in m.columns and "U_Цена" in m.columns:
        m["ΔЦена"] = pd.to_numeric(m[ex_price], errors="coerce") - pd.to_numeric(m["U_Цена"], errors="coerce")

    if ex_not in m.columns and "U_Объем" in m.columns:
        m["ΔОбъем"] = pd.to_numeric(m[ex_not], errors="coerce") - pd.to_numeric(m["U_Объем"], errors="coerce")

    m = _safe_rename(m, {
        "match_type": "Тип_совпадения",
        "key_used": "Ключ",
        "score": "Score",
    })

    if "ΔОбъем" in m.columns:
        m["_abs_diff"] = pd.to_numeric(m["ΔОбъем"], errors="coerce").abs()
        m = m.sort_values(["_abs_diff"], ascending=False).drop(columns=["_abs_diff"])

    full_front = [c for c in [
        "Тип_совпадения", "Score", "Δt_sec", "ΔQty", "ΔЦена", "ΔОбъем",
        f"{p}TradeID", f"{p}OrderID",
        f"{p}Время", f"{p}UTC", f"{p}Символ", f"{p}Сторона", f"{p}Qty", f"{p}Цена", f"{p}Объем",
        "U_ID", "U_Время", "U_UTC", "U_Instrument", "U_Символ", "U_Сторона", "U_Qty", "U_Цена", "U_Объем",
    ] if c in m.columns]
    rest = [c for c in m.columns if c not in full_front]
    m_pretty = m[full_front + rest].copy()

    if params.export_mode == "compact":
        m_pretty = m_pretty[full_front].copy()

    miss = missing_in_unity.copy()
    miss = _safe_rename(miss, {
        "Trade ID": "TradeID",
        "Order ID": "OrderID",
        "Insert Time": "Время",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "Fee": "Fee",
        "Commission Asset": "Комиссия_Asset",
    })
    miss_cols = _cols(miss, ["TradeID", "OrderID", "Время", "Символ", "Сторона", "Qty", "Цена", "Объем", "Fee", "Комиссия_Asset"])
    miss_pretty = miss[miss_cols].copy() if miss_cols else miss
    if "Символ" in miss_pretty.columns and "Время" in miss_pretty.columns:
        miss_pretty = miss_pretty.sort_values(["Символ", "Время"], ascending=[True, True])

    extra = extra_in_unity.copy()
    extra = _safe_rename(extra, {
        "ID": "ID",
        "Transact time": "Время",
        "Instrument": "Instrument",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "Net commission amount": "Комиссия",
    })
    extra_cols = _cols(extra, ["ID", "Время", "Instrument", "Символ", "Сторона", "Qty", "Цена", "Объем", "Комиссия"])
    extra_pretty = extra[extra_cols].copy() if extra_cols else extra
    if "Символ" in extra_pretty.columns and "Время" in extra_pretty.columns:
        extra_pretty = extra_pretty.sort_values(["Символ", "Время"], ascending=[True, True])

    exs = ex_status.copy()
    uns = uni_status.copy()

    u_id_map: Dict[int, Any] = unity_n["ID"].to_dict() if "ID" in unity_n.columns else {}
    if "matched_unity_idx" in exs.columns:
        def _map_uid(x: Any) -> str:
            try:
                if pd.isna(x):
                    return ""
                return str(u_id_map.get(int(x), ""))
            except Exception:
                return ""
        exs["matched_unity_id"] = exs["matched_unity_idx"].map(_map_uid)
    else:
        exs["matched_unity_id"] = ""

    exs = _safe_rename(exs, {
        "status": "Статус",
        "Trade ID": "TradeID",
        "Order ID": "OrderID",
        "Insert Time": "Время",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "matched_unity_id": "Связанный_Unity_ID",
    })
    exs_cols = _cols(exs, ["Статус", "TradeID", "OrderID", "Время", "Символ", "Сторона", "Qty", "Цена", "Объем", "Связанный_Unity_ID"])
    exs_pretty = exs[exs_cols].copy() if exs_cols else exs

    ex_tid_map: Dict[int, Any] = exchange_n["Trade ID"].to_dict() if "Trade ID" in exchange_n.columns else {}
    if "matched_exchange_idx" in uns.columns:
        def _map_etid(x: Any) -> str:
            try:
                if pd.isna(x):
                    return ""
                return str(ex_tid_map.get(int(x), ""))
            except Exception:
                return ""
        uns["matched_trade_id"] = uns["matched_exchange_idx"].map(_map_etid)
    else:
        uns["matched_trade_id"] = ""

    uns = _safe_rename(uns, {
        "status": "Статус",
        "ID": "ID",
        "Transact time": "Время",
        "Instrument": "Instrument",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "matched_trade_id": f"Связанный_{exchange_name}_TradeID",
    })
    uns_cols = _cols(uns, ["Статус", "ID", "Время", "Instrument", "Символ", "Сторона", "Qty", "Цена", "Объем", f"Связанный_{exchange_name}_TradeID"])
    uns_pretty = uns[uns_cols].copy() if uns_cols else uns

    def _vol_pretty(v: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
        if v is None:
            return None
        vv = v.copy()
        vv = _safe_rename(vv, {
            "symbol": "Символ",
            "side": "Сторона",
            "trades_exchange": f"Сделок_{exchange_name}",
            "trades_unity": "Сделок_Unity",
            "qty_sum_exchange": f"Qty_{exchange_name}",
            "qty_sum_unity": "Qty_Unity",
            "qty_diff": "ΔQty",
            "notional_sum_exchange": f"Объем_{exchange_name}",
            "notional_sum_unity": "Объем_Unity",
            "notional_diff": "ΔОбъем",
            "first_time_exchange": f"Время_первое_{exchange_name}",
            "last_time_exchange": f"Время_последнее_{exchange_name}",
            "first_time_unity": "Время_первое_Unity",
            "last_time_unity": "Время_последнее_Unity",
            "status": "Статус",
        })
        if "ΔОбъем" in vv.columns:
            vv["_absV"] = pd.to_numeric(vv["ΔОбъем"], errors="coerce").abs()
            vv = vv.sort_values(["_absV"], ascending=False).drop(columns=["_absV"])

        front_cols = _cols(vv, [
            "Статус", "Символ", "Сторона",
            f"Qty_{exchange_name}", "Qty_Unity", "ΔQty",
            f"Объем_{exchange_name}", "Объем_Unity", "ΔОбъем",
            f"Сделок_{exchange_name}", "Сделок_Unity",
        ])
        rest2 = [c for c in vv.columns if c not in front_cols]
        return vv[front_cols + rest2].copy()

    return {
        "matched": m_pretty,
        "missing": miss_pretty,
        "extra": extra_pretty,
        "ex_status": exs_pretty,
        "uni_status": uns_pretty,
        "vol_sym": _vol_pretty(volume_by_symbol),
        "vol_ss": _vol_pretty(volume_by_symbol_side),
    }


def _export_report_xlsx(
    report_path,
    summary: ReconcileSummary,
    params: ReconcileParams,
    exchange_name: str,
    matched_pretty: pd.DataFrame,
    missing_pretty: pd.DataFrame,
    extra_pretty: pd.DataFrame,
    ex_status_pretty: pd.DataFrame,
    unity_status_pretty: pd.DataFrame,
    volume_by_symbol_pretty: Optional[pd.DataFrame],
    volume_by_symbol_side_pretty: Optional[pd.DataFrame],
    top_diffs_strict: Optional[pd.DataFrame] = None,
    top_diffs_notional: Optional[pd.DataFrame] = None,
    raw_exchange: Optional[pd.DataFrame] = None,
    raw_unity: Optional[pd.DataFrame] = None,
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)

    summary_df = pd.DataFrame(
        {
            "Показатель": [
                f"Строк {exchange_name}",
                "Строк Unity",
                "Совпало STRICT",
                "Совпало FUZZY",
                "Совпало NOTIONAL (qty*price)",
                "Нет в Unity (есть в бирже)",
                "Лишнее в Unity (нет в бирже)",
                f"Объем: символов {exchange_name}",
                "Объем: символов Unity",
                "Объем: OK",
                "Объем: Расхождение",
                f"Объем: Только {exchange_name}",
                "Объем: Только Unity",
                f"Объем: Σ Qty {exchange_name}",
                "Объем: Σ Qty Unity",
                f"Объем: Σ Notional {exchange_name}",
                "Объем: Σ Notional Unity",
                f"Диапазон времени {exchange_name} (UTC)",
                "Диапазон времени Unity (UTC)",
                "Warning",
            ],
            "Значение": [
                summary.rows_exchange,
                summary.rows_unity,
                summary.matched_strict,
                summary.matched_fuzzy,
                summary.matched_notional,
                summary.missing_in_unity,
                summary.extra_in_unity,
                summary.volume_symbols_exchange,
                summary.volume_symbols_unity,
                summary.volume_symbols_ok,
                summary.volume_symbols_diff,
                summary.volume_symbols_only_exchange,
                summary.volume_symbols_only_unity,
                summary.volume_total_qty_exchange,
                summary.volume_total_qty_unity,
                summary.volume_total_notional_exchange,
                summary.volume_total_notional_unity,
                summary.exchange_time_range_utc,
                summary.unity_time_range_utc,
                summary.warning,
            ],
        }
    )

    sheet_exchange_status = f"Статус {exchange_name}"[:31]

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        _clean_df_for_excel(summary_df).to_excel(writer, sheet_name=SHEET_SUMMARY, index=False)
        _clean_df_for_excel(matched_pretty).to_excel(writer, sheet_name=SHEET_MATCHES, index=False)
        _clean_df_for_excel(missing_pretty).to_excel(writer, sheet_name=SHEET_MISSING, index=False)
        _clean_df_for_excel(extra_pretty).to_excel(writer, sheet_name=SHEET_EXTRA, index=False)
        _clean_df_for_excel(ex_status_pretty).to_excel(writer, sheet_name=sheet_exchange_status, index=False)
        _clean_df_for_excel(unity_status_pretty).to_excel(writer, sheet_name=SHEET_UNITY_STATUS, index=False)

        if volume_by_symbol_pretty is not None:
            _clean_df_for_excel(volume_by_symbol_pretty).to_excel(writer, sheet_name=SHEET_VOL_SYMBOL, index=False)
        if volume_by_symbol_side_pretty is not None:
            _clean_df_for_excel(volume_by_symbol_side_pretty).to_excel(writer, sheet_name=SHEET_VOL_SYMBOL_SIDE, index=False)

        if params.export_debug_sheets:
            if top_diffs_strict is not None:
                _clean_df_for_excel(top_diffs_strict).to_excel(writer, sheet_name="TopDiffs STRICT"[:31], index=False)
            if top_diffs_notional is not None:
                _clean_df_for_excel(top_diffs_notional).to_excel(writer, sheet_name="TopDiffs Notional"[:31], index=False)
            if raw_exchange is not None:
                _clean_df_for_excel(raw_exchange).to_excel(writer, sheet_name=f"RAW {exchange_name}"[:31], index=False)
            if raw_unity is not None:
                _clean_df_for_excel(raw_unity).to_excel(writer, sheet_name="RAW Unity"[:31], index=False)

    _style_workbook(report_path, params)


def _apply_excel_table(ws, used_names: Set[str], idx: int) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers_norm = [str(h).strip() if h is not None else "" for h in headers]
    if any(h == "" for h in headers_norm):
        return
    if len(set(headers_norm)) != len(headers_norm):
        return

    name = _make_table_name(ws.title, idx)
    j = 0
    while name in used_names:
        j += 1
        name = (name[:240] + f"_{j}")[:255]
    used_names.add(name)

    tab = Table(displayName=name, ref=ws.calculate_dimension())
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _style_workbook(report_path, params: ReconcileParams) -> None:
    wb = load_workbook(report_path)
    used_table_names: Set[str] = set()

    for i, ws in enumerate(wb.worksheets, start=1):
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        _apply_sheet_basics(ws)
        _apply_header_style(ws)
        _apply_auto_filter(ws)
        _apply_excel_table(ws, used_table_names, i)
        _apply_column_formats(ws, params)
        _apply_conditional_styles(ws)
        _autosize_columns(ws)

    wb.save(report_path)


def _apply_sheet_basics(ws) -> None:
    ws.freeze_panes = "A2" if ws.max_row >= 2 else "A1"


def _apply_header_style(ws) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=False)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22


def _apply_auto_filter(ws) -> None:
    if ws.max_row >= 2 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def _find_col_by_header(ws, header_name: str) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and v.strip() == header_name:
            return col
    return None


def _apply_column_formats(ws, params: ReconcileParams) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return

    qty_fmt = "0." + ("0" * max(0, int(params.qty_decimals)))
    price_fmt = "0." + ("0" * max(0, int(params.price_decimals)))
    notional_fmt = "0." + ("0" * max(0, int(params.notional_decimals)))
    score_fmt = "0.00"
    dt_fmt = "yyyy-mm-dd hh:mm:ss"

    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if not isinstance(header, str):
            continue
        h = header.strip()

        if ("UTC" in h) or ("Время" in h) or ("Time" in h):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = dt_fmt
            continue

        if h in {"Score", "Δt_sec"}:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = score_fmt
            continue

        if ("Qty" in h) or (h == "ΔQty"):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = qty_fmt
            continue

        if ("Цена" in h) or ("Price" in h):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = price_fmt
            continue

        if ("Объем" in h) or (h == "ΔОбъем") or ("Notional" in h) or ("Volume" in h):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = notional_fmt
            continue


def _apply_conditional_styles(ws) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return

    full_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

    fill_ok = PatternFill("solid", fgColor="E8F5E9")
    fill_warn = PatternFill("solid", fgColor="FFF8E1")
    fill_bad = PatternFill("solid", fgColor="FFEBEE")
    fill_info = PatternFill("solid", fgColor="E3F2FD")

    status_col = _find_col_by_header(ws, "Статус")
    if status_col:
        colL = get_column_letter(status_col)
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colL}2="OK"'], fill=fill_ok))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colL}2="Расхождение"'], fill=fill_bad))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,5)="Только"'], fill=fill_warn))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,7)="СОВПАЛО"'], fill=fill_ok))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,4)="НЕТ_"'], fill=fill_bad))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,7)="ЛИШНЕЕ_"'], fill=fill_warn))

    mt_col = _find_col_by_header(ws, "Тип_совпадения")
    if mt_col:
        colM = get_column_letter(mt_col)
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colM}2="STRICT"'], fill=fill_ok))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colM}2="FUZZY"'], fill=fill_warn))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colM}2="NOTIONAL"'], fill=fill_info))


def _autosize_columns(ws, sample_rows: int = 2000) -> None:
    if ws.max_column < 1:
        return
    max_r = min(ws.max_row, sample_rows)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 8
        for row in range(1, max_r + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > best:
                best = len(s)
        ws.column_dimensions[letter].width = min(max(best + 2, 10), 60)
