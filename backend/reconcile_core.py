from __future__ import annotations

import re
import uuid
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.formatting.rule import FormulaRule  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import-untyped]


# -----------------------------
# DTO / Config
# -----------------------------

@dataclass(frozen=True)
class ReconcileParams:
    # Unity transact time often contains "(UTC+5)"
    unity_utc_offset_hours: Optional[int] = 5  # None => auto-detect from text

    # BINANCE (usually UTC in exports)
    binance_delimiter: Optional[str] = ";"  # None => auto

    # OKX
    okx_utc_offset_hours: Optional[int] = None  # None => auto-detect from file metadata (if possible)
    okx_filter_trade_actions: bool = True  # keep only Buy/Sell
    okx_contract_value_overrides: Dict[str, float] = field(default_factory=dict)  # {"BTCUSDT": 0.01}
    okx_contract_value_autodetect: bool = True
    okx_contract_value_snap: bool = True
    okx_contract_value_candidates: Tuple[float, ...] = (1.0, 0.1, 0.01, 0.001, 0.0001)

    # BYBIT
    bybit_utc_offset_hours: Optional[int] = 0  # Bybit usually UTC+0 in exports
    bybit_filter_trade_actions: bool = True

    # rounding
    qty_decimals: int = 8
    price_decimals: int = 8
    notional_decimals: int = 6

    # matching
    enable_fuzzy: bool = True
    time_window_seconds: int = 180  # +- 3 min
    ignore_time_in_fuzzy: bool = False   # ✅ fuzzy без окна времени
    match_duplicates_by_time: bool = True  # ✅ strict/notional дубликаты: сортировать по времени или по исходному порядку

    qty_rel_tol: float = 1e-6
    qty_abs_tol: float = 0.0
    price_rel_tol: float = 1e-6
    price_abs_tol: float = 0.0

    enable_notional_fallback: bool = True
    notional_use_minute_bucket: bool = True  # include minute_utc in key to reduce false matches

    # volume reconciliation
    enable_volume_recon: bool = True
    volume_group_by_side: bool = True
    volume_qty_rel_tol: float = 1e-6
    volume_qty_abs_tol: float = 0.0
    volume_notional_rel_tol: float = 1e-6
    volume_notional_abs_tol: float = 0.0

    # export
    export_debug_sheets: bool = False
    export_mode: str = "compact"  # "compact" | "full"


@dataclass(frozen=True)
class ReconcileSummary:
    exchange_name: str
    rows_exchange: int
    rows_unity: int

    matched_strict: int
    matched_fuzzy: int
    matched_notional: int

    missing_in_unity: int
    extra_in_unity: int

    volume_symbols_exchange: int
    volume_symbols_unity: int
    volume_symbols_ok: int
    volume_symbols_diff: int
    volume_symbols_only_exchange: int
    volume_symbols_only_unity: int
    volume_total_qty_exchange: float
    volume_total_qty_unity: float
    volume_total_notional_exchange: float
    volume_total_notional_unity: float

    exchange_time_range_utc: str
    unity_time_range_utc: str
    warning: str = ""


@dataclass(frozen=True)
class ReconcileResult:
    report_id: str
    report_path: Path
    summary: ReconcileSummary


# -----------------------------
# Helpers (text/columns/num)
# -----------------------------

def _norm_col(s: Any) -> str:
    s2 = str(s).replace("\ufeff", "").strip()
    s2 = re.sub(r"\s+", " ", s2)
    return s2.lower()


def _pick_col(df: pd.DataFrame, candidates: List[str], required_name: str) -> str:
    norm_map = {_norm_col(c): c for c in df.columns}
    for cand in candidates:
        key = _norm_col(cand)
        if key in norm_map:
            return norm_map[key]
    raise ValueError(
        f"Не найдена колонка для '{required_name}'. "
        f"Искал: {candidates}. "
        f"В файле есть: {list(df.columns)}"
    )


def _pick_col_optional(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    norm_map = {_norm_col(c): c for c in df.columns}
    for cand in candidates:
        key = _norm_col(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def _to_numeric_series(s: pd.Series) -> pd.Series:
    if s.dtype != object:
        return pd.to_numeric(s, errors="coerce")

    t = s.astype(str).str.replace("\u00a0", "", regex=False).str.replace(" ", "", regex=False)

    has_dot = t.str.contains(r"\.", regex=True)
    has_comma = t.str.contains(",", regex=False)
    both = has_dot & has_comma

    t = t.where(~both, t.str.replace(",", "", regex=False))   # thousands "," remove
    t = t.where(both, t.str.replace(",", ".", regex=False))   # decimal comma -> dot

    return pd.to_numeric(t, errors="coerce")


def _qround_str(v: Any, decimals: int) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    d = Decimal(str(v))
    q = Decimal("1").scaleb(-decimals)
    return format(d.quantize(q, rounding=ROUND_HALF_UP), "f")


def _qround_float(v: Any, decimals: int) -> float:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return float("nan")
    d = Decimal(str(v))
    q = Decimal("1").scaleb(-decimals)
    return float(d.quantize(q, rounding=ROUND_HALF_UP))


def _detect_unity_offset_hours_from_text(s: Any, default_hours: int = 5) -> int:
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return default_hours
    txt = str(s)
    m = re.search(r"\(UTC\s*([+-])\s*(\d{1,2})(?::(\d{2}))?\)", txt, flags=re.IGNORECASE)
    if not m:
        return default_hours

    sign = 1 if m.group(1) == "+" else -1
    hh = int(m.group(2))
    mm = int(m.group(3) or "0")
    if mm != 0:
        hh = hh + (1 if mm >= 30 else 0)
    return sign * hh


def _extract_symbol_from_unity(inst: Any) -> str:
    """
    Unity Instrument examples:
      "[FU]XRPUSDT.Dec2099 :: BINA" -> XRPUSDT
      "[CFD]BTCUSDT.TOD :: OKXE"    -> BTCUSDT
      "[FU]BTCUSDT.Dec2099 :: BYBE" -> BTCUSDT
    """
    if inst is None or (isinstance(inst, float) and np.isnan(inst)):
        return ""
    s = str(inst).upper().strip()
    m = re.search(r"\](?P<sym>[A-Z0-9]+)\.", s)
    if m:
        return m.group("sym")
    m = re.search(r"\](?P<sym>[A-Z0-9]+)", s)
    if m:
        return m.group("sym")
    return re.sub(r"[^A-Z0-9]", "", s)


def _extract_symbol_from_okx(sym: Any) -> str:
    if sym is None or (isinstance(sym, float) and np.isnan(sym)):
        return ""
    s = str(sym).upper().strip().replace(" ", "")
    parts = s.split("-")
    if len(parts) >= 2:
        return f"{parts[0]}{parts[1]}"
    return re.sub(r"[^A-Z0-9]", "", s)


def _extract_symbol_basic(sym: Any) -> str:
    if sym is None or (isinstance(sym, float) and np.isnan(sym)):
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(sym).upper().strip())


def _snap_value(x: float, candidates: Tuple[float, ...], rel_tol: float = 0.05) -> Optional[float]:
    if not np.isfinite(x) or x <= 0:
        return None
    best = None
    best_err = None
    for c in candidates:
        err = abs(x - c) / c
        if best_err is None or err < best_err:
            best_err = err
            best = c
    if best is None:
        return None
    if best_err is not None and best_err <= rel_tol:
        return float(best)
    return float(x)


# -----------------------------
# Readers
# -----------------------------

def _read_delimited_text(path: Path, delimiter: Optional[str]) -> pd.DataFrame:
    if delimiter:
        df = pd.read_csv(path, sep=delimiter, engine="python")
        if df.shape[1] > 1:
            return df
    for sep in [";", ",", "\t", "|"]:
        try:
            df = pd.read_csv(path, sep=sep, engine="python")
            if df.shape[1] > 1:
                return df
        except Exception:
            pass
    return pd.read_csv(path, engine="python")


def _read_binance_file(path: Path, delimiter: Optional[str]) -> pd.DataFrame:
    suf = path.suffix.lower()
    if suf in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
        df.columns = [str(c).replace("\ufeff", "").strip() for c in df.columns]
        return df
    return _read_delimited_text(path, delimiter)


def _read_bybit_file(path: Path, delimiter: Optional[str] = None) -> pd.DataFrame:
    suf = path.suffix.lower()
    if suf in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
        df.columns = [str(c).replace("\ufeff", "").strip() for c in df.columns]
        return df
    return _read_delimited_text(path, delimiter)


def _detect_okx_header_row(path: Path, max_rows: int = 25) -> int:
    try:
        raw = pd.read_excel(path, header=None, nrows=max_rows)
    except Exception:
        return 1

    def row_text(vals: List[Any]) -> str:
        return " | ".join([str(v) for v in vals if v is not None]).lower()

    for i in range(min(max_rows, len(raw))):
        s = row_text(raw.iloc[i].tolist())
        has_time = "time" in s
        has_symbol = ("symbol" in s) or ("instrument" in s) or ("inst" in s)
        has_action = ("action" in s) or ("side" in s) or ("direction" in s)
        if has_time and has_symbol and has_action:
            return i

    return 1


def _read_okx_xlsx(path: Path) -> Tuple[pd.DataFrame, Optional[int]]:
    tz_offset = None
    try:
        head = pd.read_excel(path, header=None, nrows=5)
        for v in head.values.flatten().tolist():
            if v is None:
                continue
            txt = str(v)
            m = re.search(r"UTC\s*([+-])\s*(\d{1,2})", txt, flags=re.IGNORECASE)
            if m:
                sign = 1 if m.group(1) == "+" else -1
                tz_offset = sign * int(m.group(2))
                break
    except Exception:
        tz_offset = None

    header_row = _detect_okx_header_row(path)
    df = pd.read_excel(path, header=header_row)
    df.columns = [str(c).replace("\ufeff", "").strip() for c in df.columns]
    return df, tz_offset


# -----------------------------
# Exchange standardizers
# -----------------------------

def _prepare_okx_to_standard(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).replace("\ufeff", "").strip() for c in out.columns]

    col_time = _pick_col(out, ["Time", "Trade Time", "Fill Time", "Timestamp"], "Insert Time")
    col_symbol = _pick_col(out, ["Symbol", "Instrument", "Inst", "Trading Pair", "Currency Pair"], "Symbol")
    col_side = _pick_col(out, ["Action", "Side", "Direction"], "Side")
    col_qty = _pick_col(out, ["Amount", "Quantity", "Size", "Filled Amount", "Qty"], "Quantity")
    col_price = _pick_col(out, ["Filled Price", "Price", "Fill Price", "Avg Price"], "Price")

    col_trade_id = _pick_col_optional(out, ["id", "Trade ID", "Fill ID", "trade id", "﻿id"])
    col_order_id = _pick_col_optional(out, ["Order id", "Order ID", "order id"])
    col_fee = _pick_col_optional(out, ["Fee", "Trading Fee", "Commission"])
    col_fee_unit = _pick_col_optional(out, ["Fee Unit", "Commission Asset", "Fee Currency", "Fee Coin"])
    col_unit = _pick_col_optional(out, ["Trading Unit", "Unit"])

    std = pd.DataFrame()
    std["Insert Time"] = out[col_time]
    std["Symbol"] = out[col_symbol]
    std["Side"] = out[col_side]
    std["Quantity"] = out[col_qty]
    std["Price"] = out[col_price]

    if col_trade_id:
        std["Trade ID"] = out[col_trade_id]
    if col_order_id:
        std["Order ID"] = out[col_order_id]
    if col_fee:
        std["Fee"] = out[col_fee]
    if col_fee_unit:
        std["Commission Asset"] = out[col_fee_unit]
    if col_unit:
        std["Trading Unit"] = out[col_unit]

    return std


def _prepare_binance_to_standard(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).replace("\ufeff", "").strip() for c in out.columns]

    col_time = _pick_col(out, ["Insert Time", "Time", "Date(UTC)", "Date (UTC)", "Trade Time"], "Insert Time")
    col_symbol = _pick_col(out, ["Symbol", "Pair", "Market", "Instrument"], "Symbol")
    col_side = _pick_col(out, ["Side", "Action", "Direction"], "Side")
    col_price = _pick_col(out, ["Price", "Avg Price", "Filled Price"], "Price")

    col_qty = None
    for cand in ["Quantity", "Qty", "Executed", "Amount", "Filled Amount"]:
        try:
            col_qty = _pick_col(out, [cand], "Quantity")
            break
        except Exception:
            pass
    if not col_qty:
        raise ValueError(f"Не найдена колонка количества (Quantity/Qty/Executed/Amount). Колонки: {list(out.columns)}")

    col_trade_id = _pick_col_optional(out, ["Trade ID", "TradeId", "ID", "id"])
    col_order_id = _pick_col_optional(out, ["Order ID", "OrderId", "Order id", "Order No.", "Order No"])
    col_fee = _pick_col_optional(out, ["Fee", "Commission", "Trading Fee"])
    col_fee_asset = _pick_col_optional(out, ["Commission Asset", "Fee Unit", "Fee Asset", "Commission Coin"])

    std = pd.DataFrame()
    std["Insert Time"] = out[col_time]
    std["Symbol"] = out[col_symbol]
    std["Side"] = out[col_side]
    std["Quantity"] = out[col_qty]
    std["Price"] = out[col_price]

    if col_trade_id:
        std["Trade ID"] = out[col_trade_id]
    if col_order_id:
        std["Order ID"] = out[col_order_id]
    if col_fee:
        std["Fee"] = out[col_fee]
    if col_fee_asset:
        std["Commission Asset"] = out[col_fee_asset]

    return std


def _map_bybit_side(v: Any) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    s = str(v).strip().upper()

    if s in {"BUY", "B"}:
        return "BUY"
    if s in {"SELL", "S"}:
        return "SELL"

    if s == "LONG":
        return "BUY"
    if s == "SHORT":
        return "SELL"

    if "CLOSE" in s and "LONG" in s:
        return "SELL"
    if "CLOSE" in s and "SHORT" in s:
        return "BUY"

    if "BUY" in s:
        return "BUY"
    if "SELL" in s:
        return "SELL"
    if "LONG" in s:
        return "BUY"
    if "SHORT" in s:
        return "SELL"
    return s


def _prepare_bybit_to_standard(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).replace("\ufeff", "").strip() for c in out.columns]

    col_time = _pick_col(
        out,
        [
            "Transaction Time(UTC+0)",
            "Transaction Time (UTC+0)",
            "Transaction Time",
            "Time",
            "Date(UTC)",
            "Date (UTC)",
        ],
        "Insert Time",
    )
    col_symbol = _pick_col(out, ["Market", "Symbol", "Pair", "Instrument"], "Symbol")
    col_side = _pick_col(out, ["Direction", "Side", "Action"], "Side")
    col_qty = _pick_col(out, ["Filled Quantity", "Qty", "Quantity", "Executed", "Amount", "Size"], "Quantity")
    col_price = _pick_col(out, ["Filled Price", "Price", "Avg Price", "Executed Price"], "Price")

    col_trade_id = _pick_col_optional(out, ["Transaction ID", "Trasaction ID", "Trade ID", "Exec ID", "Fill ID", "ID", "id"])
    col_order_id = _pick_col_optional(out, ["Order No.", "Order No", "Order ID", "OrderId", "Order id"])
    col_fee = _pick_col_optional(out, ["Trading Fee", "Fee", "Commission", "ExecFeeV2", "Exec Fee", "ExecFee"])
    col_fee_asset = _pick_col_optional(out, ["feeCoin", "Fee Coin", "Commission Asset", "Fee Unit", "Fee Asset"])

    std = pd.DataFrame()
    std["Insert Time"] = out[col_time]
    std["Symbol"] = out[col_symbol]
    std["Side"] = out[col_side].map(_map_bybit_side)
    std["Quantity"] = out[col_qty]
    std["Price"] = out[col_price]

    if col_trade_id:
        std["Trade ID"] = out[col_trade_id]
    if col_order_id:
        std["Order ID"] = out[col_order_id]
    if col_fee:
        std["Fee"] = out[col_fee]
    if col_fee_asset:
        std["Commission Asset"] = out[col_fee_asset]

    return std


# -----------------------------
# Normalization
# -----------------------------

def _normalize_unity(df: pd.DataFrame, params: ReconcileParams) -> Tuple[pd.DataFrame, int]:
    need_cols = ["Instrument", "Side", "Transact time", "Price"]
    for c in need_cols:
        if c not in df.columns:
            raise ValueError(f"Unity file missing required column: {c}")

    out = df.copy()
    out["symbol"] = out["Instrument"].map(_extract_symbol_from_unity)
    out["side"] = out["Side"].astype(str).str.strip().str.upper()

    qty_col = "Absolute amount" if "Absolute amount" in out.columns else ("Amount" if "Amount" in out.columns else None)
    if not qty_col:
        raise ValueError("Unity file must contain 'Absolute amount' or 'Amount'")

    out["qty"] = _to_numeric_series(out[qty_col]).abs()
    out["price"] = _to_numeric_series(out["Price"])

    offset = params.unity_utc_offset_hours
    if offset is None:
        sample = out["Transact time"].dropna().astype(str).head(20).tolist()
        offset = _detect_unity_offset_hours_from_text(sample[0], default_hours=5) if sample else 5

    tt = out["Transact time"].astype(str).str.replace(r"\s*\(UTC[^\)]*\)\s*", "", regex=True).str.strip()
    out["trade_dt_local"] = pd.to_datetime(tt, dayfirst=True, errors="coerce")
    out["trade_dt_utc"] = out["trade_dt_local"] - pd.Timedelta(hours=int(offset))
    out["minute_utc"] = out["trade_dt_utc"].dt.floor("min")

    out["qty_r"] = out["qty"].map(lambda x: _qround_str(x, params.qty_decimals))
    out["price_r"] = out["price"].map(lambda x: _qround_str(x, params.price_decimals))

    out["match_key"] = (
        out["symbol"].astype(str) + "|" +
        out["side"].astype(str) + "|" +
        out["qty_r"].astype(str) + "|" +
        out["price_r"].astype(str)
    )

    out["notional"] = out["qty"] * out["price"]
    out["notional_r"] = out["notional"].map(lambda x: _qround_str(x, params.notional_decimals))

    if params.notional_use_minute_bucket:
        out["notional_key"] = (
            out["symbol"].astype(str) + "|" +
            out["side"].astype(str) + "|" +
            out["minute_utc"].astype(str) + "|" +
            out["notional_r"].astype(str)
        )
    else:
        out["notional_key"] = (
            out["symbol"].astype(str) + "|" +
            out["side"].astype(str) + "|" +
            out["notional_r"].astype(str)
        )

    return out, int(offset)


def _normalize_exchange_common(
    df: pd.DataFrame,
    params: ReconcileParams,
    *,
    time_offset_hours: int = 0,
    symbol_mapper: Callable[[Any], str],
    action_filter: Optional[set[str]] = None,
    contract_value_map: Optional[Dict[str, float]] = None,
    trading_unit_col: Optional[str] = None,
) -> pd.DataFrame:
    need_cols = ["Symbol", "Side", "Quantity", "Price", "Insert Time"]
    for c in need_cols:
        if c not in df.columns:
            raise ValueError(f"Exchange file missing required column: {c}. Available: {list(df.columns)}")

    out = df.copy()

    out["Symbol"] = out["Symbol"].map(symbol_mapper)
    out["Side"] = out["Side"].astype(str).str.strip().str.upper()

    qty_raw = _to_numeric_series(out["Quantity"]).abs()
    price = _to_numeric_series(out["Price"])

    if action_filter is not None:
        mask = out["Side"].isin(action_filter)
        out = out.loc[mask].copy()
        qty_raw = qty_raw.loc[mask]
        price = price.loc[mask]

    out["symbol"] = out["Symbol"].astype(str).str.strip().str.upper()
    out["side"] = out["Side"].astype(str).str.strip().str.upper()

    mult = pd.Series(1.0, index=out.index)
    if contract_value_map:
        mult = out["symbol"].map(lambda s: float(contract_value_map.get(str(s), 1.0))).astype(float)

    if trading_unit_col and trading_unit_col in out.columns:
        unit = out[trading_unit_col].astype(str).str.strip().str.lower()
        is_cont = unit.isin(["cont", "contract", "contracts"])
        out["qty"] = qty_raw * np.where(is_cont, mult, 1.0)
    else:
        out["qty"] = qty_raw * mult

    out["price"] = price

    out["trade_dt_local"] = pd.to_datetime(out["Insert Time"], dayfirst=True, errors="coerce")
    out["trade_dt_utc"] = out["trade_dt_local"] - pd.Timedelta(hours=int(time_offset_hours))
    out["minute_utc"] = out["trade_dt_utc"].dt.floor("min")

    out["qty_r"] = out["qty"].map(lambda x: _qround_str(x, params.qty_decimals))
    out["price_r"] = out["price"].map(lambda x: _qround_str(x, params.price_decimals))

    out["match_key"] = (
        out["symbol"].astype(str) + "|" +
        out["side"].astype(str) + "|" +
        out["qty_r"].astype(str) + "|" +
        out["price_r"].astype(str)
    )

    out["notional"] = out["qty"] * out["price"]
    out["notional_r"] = out["notional"].map(lambda x: _qround_str(x, params.notional_decimals))

    if params.notional_use_minute_bucket:
        out["notional_key"] = (
            out["symbol"].astype(str) + "|" +
            out["side"].astype(str) + "|" +
            out["minute_utc"].astype(str) + "|" +
            out["notional_r"].astype(str)
        )
    else:
        out["notional_key"] = (
            out["symbol"].astype(str) + "|" +
            out["side"].astype(str) + "|" +
            out["notional_r"].astype(str)
        )

    return out


def _infer_okx_contract_value_map(
    unity_n: pd.DataFrame,
    okx_std: pd.DataFrame,
    symbol_mapper: Callable[[Any], str],
    action_filter: Optional[set[str]],
    params: ReconcileParams,
) -> Dict[str, float]:
    if not params.okx_contract_value_autodetect:
        return dict(params.okx_contract_value_overrides)

    if "Quantity" not in okx_std.columns or "Symbol" not in okx_std.columns:
        return dict(params.okx_contract_value_overrides)

    tmp = okx_std.copy()
    tmp["Symbol"] = tmp["Symbol"].map(symbol_mapper)
    tmp["Side"] = tmp["Side"].astype(str).str.strip().str.upper()
    if action_filter is not None:
        tmp = tmp[tmp["Side"].isin(action_filter)].copy()

    tmp["symbol"] = tmp["Symbol"].astype(str).str.strip().str.upper()
    qty_contracts = _to_numeric_series(tmp["Quantity"]).abs()
    e_tot = qty_contracts.groupby(tmp["symbol"]).sum()
    u_tot = unity_n.groupby("symbol")["qty"].sum()

    out: Dict[str, float] = dict(params.okx_contract_value_overrides)
    for sym in (set(u_tot.index) & set(e_tot.index)):
        if sym in out:
            continue
        denom = float(e_tot.loc[sym])
        num = float(u_tot.loc[sym])
        if denom <= 0 or num <= 0:
            continue
        ratio = num / denom
        if params.okx_contract_value_snap:
            snapped = _snap_value(ratio, params.okx_contract_value_candidates, rel_tol=0.10)
            out[sym] = float(snapped if snapped is not None else ratio)
        else:
            out[sym] = float(ratio)

    return out


# -----------------------------
# Matching
# -----------------------------

def _within_tol(bv: float, uv: float, rel: float, abs_tol: float) -> bool:
    if pd.isna(bv) or pd.isna(uv):
        return False
    diff = abs(float(uv) - float(bv))
    thresh = max(abs_tol, rel * abs(float(bv)))
    return diff <= thresh


def _group_indices(df: pd.DataFrame, key_col: str, *, sort_by_time: bool) -> Dict[Any, List[int]]:
    if df.empty:
        return {}
    if sort_by_time and "trade_dt_utc" in df.columns:
        df_sorted = df.sort_values("trade_dt_utc", kind="mergesort")
    else:
        df_sorted = df.copy()

    idx_map = df_sorted.groupby(key_col, sort=False).indices
    res: Dict[Any, List[int]] = {}
    idx_values = df_sorted.index.to_numpy()
    for k, pos in idx_map.items():
        res[k] = idx_values[pos].tolist()
    return res


def _reconcile_multiset_by_key(
    unity_n: pd.DataFrame,
    exchange_n: pd.DataFrame,
    key_col: str,
    *,
    sort_by_time: bool,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    u_groups = _group_indices(unity_n, key_col, sort_by_time=sort_by_time)
    e_groups = _group_indices(exchange_n, key_col, sort_by_time=sort_by_time)

    matched: List[Tuple[int, int, str]] = []
    missing_ex: List[int] = []
    extra_unity: List[int] = []

    for k, e_idxs in e_groups.items():
        u_idxs = u_groups.get(k, [])
        n = min(len(e_idxs), len(u_idxs))
        for i in range(n):
            matched.append((int(e_idxs[i]), int(u_idxs[i]), str(k)))
        if len(e_idxs) > n:
            missing_ex.extend([int(x) for x in e_idxs[n:]])
        if len(u_idxs) > n:
            extra_unity.extend([int(x) for x in u_idxs[n:]])

    for k, u_idxs in u_groups.items():
        if k not in e_groups:
            extra_unity.extend([int(x) for x in u_idxs])

    matched_df = pd.DataFrame(matched, columns=["exchange_idx", "unity_idx", "key"])
    missing_df = exchange_n.loc[missing_ex].copy()
    extra_df = unity_n.loc[extra_unity].copy()
    return matched_df, missing_df, extra_df


def _reconcile_fuzzy(
    missing_exchange: pd.DataFrame,
    extra_unity: pd.DataFrame,
    params: ReconcileParams,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if missing_exchange.empty or extra_unity.empty:
        return (
            pd.DataFrame(columns=["exchange_idx", "unity_idx", "score"]),
            missing_exchange,
            extra_unity,
        )

    b = missing_exchange.copy()
    u = extra_unity.copy()

    b["_idx"] = b.index.astype(int)
    u["_idx"] = u.index.astype(int)

    b["grp"] = b["symbol"].astype(str) + "|" + b["side"].astype(str)
    u["grp"] = u["symbol"].astype(str) + "|" + u["side"].astype(str)

    # sort only if time is used
    if (not params.ignore_time_in_fuzzy) and ("trade_dt_utc" in b.columns):
        b = b.sort_values("trade_dt_utc", kind="mergesort")
    if (not params.ignore_time_in_fuzzy) and ("trade_dt_utc" in u.columns):
        u = u.sort_values("trade_dt_utc", kind="mergesort")

    used_unity: set[int] = set()
    matched_rows: List[Dict[str, Any]] = []

    groups: Dict[str, Dict[str, Any]] = {}
    for grp, g in u.groupby("grp", sort=False):
        groups[str(grp)] = {
            "ids": g["_idx"].to_numpy(dtype=np.int64),
            "qty": pd.to_numeric(g["qty"], errors="coerce").to_numpy(dtype=float),
            "price": pd.to_numeric(g["price"], errors="coerce").to_numpy(dtype=float),
            "time": g["trade_dt_utc"].to_numpy(dtype="datetime64[ns]") if ("trade_dt_utc" in g.columns) else None,
        }

    win = pd.Timedelta(seconds=int(params.time_window_seconds))
    win_ns = np.int64(win.value)

    for _, brow in b.iterrows():
        grp = str(brow["grp"])
        if grp not in groups:
            continue

        b_qty = float(brow["qty"]) if not pd.isna(brow["qty"]) else np.nan
        b_price = float(brow["price"]) if not pd.isna(brow["price"]) else np.nan
        if not np.isfinite(b_qty) or not np.isfinite(b_price):
            continue

        g = groups[grp]
        ids_u = g["ids"]
        q_u = g["qty"]
        p_u = g["price"]

        qty_thr = max(params.qty_abs_tol, params.qty_rel_tol * abs(b_qty))
        price_thr = max(params.price_abs_tol, params.price_rel_tol * abs(b_price))

        mask = np.isfinite(q_u) & np.isfinite(p_u)
        mask &= (np.abs(q_u - b_qty) <= qty_thr)
        mask &= (np.abs(p_u - b_price) <= price_thr)

        if not params.ignore_time_in_fuzzy:
            bt = brow.get("trade_dt_utc", pd.NaT)
            if pd.isna(bt):
                continue
            bt64 = np.datetime64(bt.to_datetime64())
            t_u = g["time"]
            if t_u is not None:
                mask &= (t_u >= bt64 - np.timedelta64(win_ns, "ns"))
                mask &= (t_u <= bt64 + np.timedelta64(win_ns, "ns"))

        cand_pos = np.where(mask)[0]
        if cand_pos.size == 0:
            continue

        best_score = None
        best_uid = None

        for pos in cand_pos:
            uid = int(ids_u[pos])
            if uid in used_unity:
                continue

            qty_rel = abs(q_u[pos] - b_qty) / (abs(b_qty) if abs(b_qty) > 0 else 1.0)
            price_rel = abs(p_u[pos] - b_price) / (abs(b_price) if abs(b_price) > 0 else 1.0)

            if params.ignore_time_in_fuzzy:
                score = 1000.0 * qty_rel + 1000.0 * price_rel
            else:
                bt = brow["trade_dt_utc"]
                ut = pd.Timestamp(g["time"][pos])
                dt_sec = abs((ut.to_pydatetime() - bt).total_seconds())
                score = dt_sec + 1000.0 * qty_rel + 1000.0 * price_rel

            if best_score is None or score < best_score:
                best_score = score
                best_uid = uid

        if best_uid is not None and best_score is not None:
            used_unity.add(best_uid)
            matched_rows.append({"exchange_idx": int(brow["_idx"]), "unity_idx": int(best_uid), "score": float(best_score)})

    matched_fuzzy = pd.DataFrame(matched_rows, columns=["exchange_idx", "unity_idx", "score"])

    if not matched_fuzzy.empty:
        matched_b = set(matched_fuzzy["exchange_idx"].tolist())
        matched_u = set(matched_fuzzy["unity_idx"].tolist())
        missing_after = missing_exchange[~missing_exchange.index.isin(matched_b)].copy()
        extra_after = extra_unity[~extra_unity.index.isin(matched_u)].copy()
    else:
        missing_after = missing_exchange
        extra_after = extra_unity

    return matched_fuzzy, missing_after, extra_after


# -----------------------------
# Volume reconciliation
# -----------------------------

def _agg_volume(df: pd.DataFrame, by_side: bool) -> pd.DataFrame:
    keys = ["symbol"] + (["side"] if by_side else [])
    g = (
        df.groupby(keys, dropna=False)
        .agg(
            trades=("symbol", "size"),
            qty_sum=("qty", "sum"),
            notional_sum=("notional", "sum"),
            first_time=("trade_dt_utc", "min"),
            last_time=("trade_dt_utc", "max"),
        )
        .reset_index()
    )
    return g


def _compare_volume(agg_ex: pd.DataFrame, agg_u: pd.DataFrame, by_side: bool, params: ReconcileParams) -> pd.DataFrame:
    keys = ["symbol"] + (["side"] if by_side else [])
    m = agg_ex.merge(agg_u, on=keys, how="outer", suffixes=("_exchange", "_unity"))

    for col in ["trades", "qty_sum", "notional_sum"]:
        ce = f"{col}_exchange"
        cu = f"{col}_unity"
        if ce not in m.columns:
            m[ce] = 0
        if cu not in m.columns:
            m[cu] = 0
        m[ce] = m[ce].fillna(0)
        m[cu] = m[cu].fillna(0)

    m["qty_diff"] = m["qty_sum_unity"] - m["qty_sum_exchange"]
    m["notional_diff"] = m["notional_sum_unity"] - m["notional_sum_exchange"]

    statuses: List[str] = []
    for tb, tu, qb, qu, nb, nu in zip(
        m["trades_exchange"], m["trades_unity"],
        m["qty_sum_exchange"], m["qty_sum_unity"],
        m["notional_sum_exchange"], m["notional_sum_unity"],
    ):
        if tb == 0 and tu > 0:
            statuses.append("Только Unity")
            continue
        if tu == 0 and tb > 0:
            statuses.append("Только Биржа")
            continue

        qty_ok = _within_tol(float(qb), float(qu), params.volume_qty_rel_tol, params.volume_qty_abs_tol)
        not_ok = _within_tol(float(nb), float(nu), params.volume_notional_rel_tol, params.volume_notional_abs_tol)
        statuses.append("OK" if (qty_ok and not_ok) else "Расхождение")

    m["status"] = statuses

    m["qty_sum_exchange"] = m["qty_sum_exchange"].map(lambda x: _qround_float(x, params.qty_decimals))
    m["qty_sum_unity"] = m["qty_sum_unity"].map(lambda x: _qround_float(x, params.qty_decimals))
    m["qty_diff"] = m["qty_diff"].map(lambda x: _qround_float(x, params.qty_decimals))

    m["notional_sum_exchange"] = m["notional_sum_exchange"].map(lambda x: _qround_float(x, params.notional_decimals))
    m["notional_sum_unity"] = m["notional_sum_unity"].map(lambda x: _qround_float(x, params.notional_decimals))
    m["notional_diff"] = m["notional_diff"].map(lambda x: _qround_float(x, params.notional_decimals))

    m["_abs_not"] = pd.to_numeric(m["notional_diff"], errors="coerce").abs()
    m = m.sort_values("_abs_not", ascending=False).drop(columns=["_abs_not"])
    return m


# -----------------------------
# Debug helpers
# -----------------------------

def _top_key_diffs(unity_n: pd.DataFrame, exchange_n: pd.DataFrame, key_col: str, limit: int = 50) -> pd.DataFrame:
    u_cnt = unity_n[key_col].value_counts()
    e_cnt = exchange_n[key_col].value_counts()
    diff = (u_cnt.subtract(e_cnt, fill_value=0)).sort_values(key=lambda s: s.abs(), ascending=False)
    out = diff.head(limit).reset_index()
    out.columns = [key_col, "unity_count_minus_exchange_count"]
    return out


# -----------------------------
# Pretty report helpers
# -----------------------------

def _cols(df: pd.DataFrame, ordered: List[str]) -> List[str]:
    return [c for c in ordered if c in df.columns]


def _safe_rename(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    m = {k: v for k, v in mapping.items() if k in df.columns}
    return df.rename(columns=m)


def _build_pretty_tables(
    matched_all: pd.DataFrame,
    exchange_n: pd.DataFrame,
    unity_n: pd.DataFrame,
    missing_in_unity: pd.DataFrame,
    extra_in_unity: pd.DataFrame,
    ex_status: pd.DataFrame,
    uni_status: pd.DataFrame,
    volume_by_symbol: Optional[pd.DataFrame],
    volume_by_symbol_side: Optional[pd.DataFrame],
    *,
    exchange_name: str,
    exchange_code: str,
    params: ReconcileParams,
) -> Dict[str, Optional[pd.DataFrame]]:
    p = f"{exchange_code}_"

    ex_sel = exchange_n.reset_index().rename(columns={"index": "exchange_idx"})
    u_sel = unity_n.reset_index().rename(columns={"index": "unity_idx"})

    ex_keep = _cols(ex_sel, [
        "exchange_idx",
        "Trade ID", "Order ID", "Insert Time",
        "trade_dt_utc",
        "symbol", "side", "qty", "price", "notional",
        "Fee", "Commission Asset",
    ])
    u_keep = _cols(u_sel, [
        "unity_idx",
        "ID", "Transact time",
        "trade_dt_utc",
        "Instrument",
        "symbol", "side", "qty", "price", "notional",
        "Net commission amount",
    ])

    ex_view = ex_sel[ex_keep].copy()
    u_view = u_sel[u_keep].copy()

    ex_view = _safe_rename(ex_view, {
        "Trade ID": f"{p}TradeID",
        "Order ID": f"{p}OrderID",
        "Insert Time": f"{p}Время",
        "trade_dt_utc": f"{p}UTC",
        "symbol": f"{p}Символ",
        "side": f"{p}Сторона",
        "qty": f"{p}Qty",
        "price": f"{p}Цена",
        "notional": f"{p}Объем",
        "Fee": f"{p}Fee",
        "Commission Asset": f"{p}Комиссия_Asset",
    })

    u_view = _safe_rename(u_view, {
        "ID": "U_ID",
        "Transact time": "U_Время",
        "trade_dt_utc": "U_UTC",
        "Instrument": "U_Instrument",
        "symbol": "U_Символ",
        "side": "U_Сторона",
        "qty": "U_Qty",
        "price": "U_Цена",
        "notional": "U_Объем",
        "Net commission amount": "U_Комиссия",
    })

    m = matched_all.copy()
    if "score" not in m.columns:
        m["score"] = np.nan
    if "key_used" not in m.columns:
        m["key_used"] = ""

    m = m.merge(ex_view, on="exchange_idx", how="left").merge(u_view, on="unity_idx", how="left")

    ex_utc = f"{p}UTC"
    ex_qty = f"{p}Qty"
    ex_price = f"{p}Цена"
    ex_not = f"{p}Объем"

    if ex_utc in m.columns and "U_UTC" in m.columns:
        t1 = pd.to_datetime(m[ex_utc], errors="coerce")
        t2 = pd.to_datetime(m["U_UTC"], errors="coerce")
        m["Δt_sec"] = (t1 - t2).abs().dt.total_seconds()

    if ex_qty in m.columns and "U_Qty" in m.columns:
        m["ΔQty"] = pd.to_numeric(m[ex_qty], errors="coerce") - pd.to_numeric(m["U_Qty"], errors="coerce")
    if ex_price in m.columns and "U_Цена" in m.columns:
        m["ΔЦена"] = pd.to_numeric(m[ex_price], errors="coerce") - pd.to_numeric(m["U_Цена"], errors="coerce")
    if ex_not in m.columns and "U_Объем" in m.columns:
        m["ΔОбъем"] = pd.to_numeric(m[ex_not], errors="coerce") - pd.to_numeric(m["U_Объем"], errors="coerce")

    m = _safe_rename(m, {"match_type": "Тип_совпадения", "key_used": "Ключ", "score": "Score"})

    if "ΔОбъем" in m.columns:
        m["_abs_diff"] = pd.to_numeric(m["ΔОбъем"], errors="coerce").abs()
        m = m.sort_values(["_abs_diff"], ascending=False).drop(columns=["_abs_diff"])

    full_front = [c for c in [
        "Тип_совпадения", "Score", "Δt_sec", "ΔQty", "ΔЦена", "ΔОбъем",
        f"{p}TradeID", f"{p}OrderID",
        f"{p}Время", f"{p}UTC", f"{p}Символ", f"{p}Сторона", f"{p}Qty", f"{p}Цена", f"{p}Объем",
        "U_ID", "U_Время", "U_UTC", "U_Instrument", "U_Символ", "U_Сторона", "U_Qty", "U_Цена", "U_Объем",
    ] if c in m.columns]
    rest = [c for c in m.columns if c not in full_front]
    m_pretty = m[full_front + rest].copy()
    if params.export_mode == "compact":
        m_pretty = m_pretty[full_front].copy()

    miss = missing_in_unity.copy()
    miss = _safe_rename(miss, {
        "Trade ID": "TradeID",
        "Order ID": "OrderID",
        "Insert Time": "Время",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "Fee": "Fee",
        "Commission Asset": "Комиссия_Asset",
    })
    miss_cols = _cols(miss, ["TradeID", "OrderID", "Время", "Символ", "Сторона", "Qty", "Цена", "Объем", "Fee", "Комиссия_Asset"])
    miss_pretty = miss[miss_cols].copy() if miss_cols else miss

    extra = extra_in_unity.copy()
    extra = _safe_rename(extra, {
        "ID": "ID",
        "Transact time": "Время",
        "Instrument": "Instrument",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "Net commission amount": "Комиссия",
    })
    extra_cols = _cols(extra, ["ID", "Время", "Instrument", "Символ", "Сторона", "Qty", "Цена", "Объем", "Комиссия"])
    extra_pretty = extra[extra_cols].copy() if extra_cols else extra

    exs = ex_status.copy()
    uns = uni_status.copy()

    u_id_map: Dict[int, Any] = unity_n["ID"].to_dict() if "ID" in unity_n.columns else {}
    if "matched_unity_idx" in exs.columns:
        def _map_uid(x: Any) -> str:
            try:
                if pd.isna(x):
                    return ""
                return str(u_id_map.get(int(x), ""))
            except Exception:
                return ""
        exs["matched_unity_id"] = exs["matched_unity_idx"].map(_map_uid)
    else:
        exs["matched_unity_id"] = ""

    exs = _safe_rename(exs, {
        "status": "Статус",
        "Trade ID": "TradeID",
        "Order ID": "OrderID",
        "Insert Time": "Время",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "matched_unity_id": "Связанный_Unity_ID",
    })
    exs_cols = _cols(exs, ["Статус", "TradeID", "OrderID", "Время", "Символ", "Сторона", "Qty", "Цена", "Объем", "Связанный_Unity_ID"])
    exs_pretty = exs[exs_cols].copy() if exs_cols else exs

    ex_tid_map: Dict[int, Any] = exchange_n["Trade ID"].to_dict() if "Trade ID" in exchange_n.columns else {}
    if "matched_exchange_idx" in uns.columns:
        def _map_etid(x: Any) -> str:
            try:
                if pd.isna(x):
                    return ""
                return str(ex_tid_map.get(int(x), ""))
            except Exception:
                return ""
        uns["matched_trade_id"] = uns["matched_exchange_idx"].map(_map_etid)
    else:
        uns["matched_trade_id"] = ""

    uns = _safe_rename(uns, {
        "status": "Статус",
        "ID": "ID",
        "Transact time": "Время",
        "Instrument": "Instrument",
        "symbol": "Символ",
        "side": "Сторона",
        "qty": "Qty",
        "price": "Цена",
        "notional": "Объем",
        "matched_trade_id": f"Связанный_{exchange_name}_TradeID",
    })
    uns_cols = _cols(uns, ["Статус", "ID", "Время", "Instrument", "Символ", "Сторона", "Qty", "Цена", "Объем", f"Связанный_{exchange_name}_TradeID"])
    uns_pretty = uns[uns_cols].copy() if uns_cols else uns

    def _vol_pretty(v: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
        if v is None:
            return None
        vv = v.copy()
        vv = _safe_rename(vv, {
            "symbol": "Символ",
            "side": "Сторона",
            "trades_exchange": f"Сделок_{exchange_name}",
            "trades_unity": "Сделок_Unity",
            "qty_sum_exchange": f"Qty_{exchange_name}",
            "qty_sum_unity": "Qty_Unity",
            "qty_diff": "ΔQty",
            "notional_sum_exchange": f"Объем_{exchange_name}",
            "notional_sum_unity": "Объем_Unity",
            "notional_diff": "ΔОбъем",
            "first_time_exchange": f"Время_первое_{exchange_name}",
            "last_time_exchange": f"Время_последнее_{exchange_name}",
            "first_time_unity": "Время_первое_Unity",
            "last_time_unity": "Время_последнее_Unity",
            "status": "Статус",
        })
        if "ΔОбъем" in vv.columns:
            vv["_absV"] = pd.to_numeric(vv["ΔОбъем"], errors="coerce").abs()
            vv = vv.sort_values(["_absV"], ascending=False).drop(columns=["_absV"])
        return vv

    return {
        "matched": m_pretty,
        "missing": miss_pretty,
        "extra": extra_pretty,
        "ex_status": exs_pretty,
        "uni_status": uns_pretty,
        "vol_sym": _vol_pretty(volume_by_symbol),
        "vol_ss": _vol_pretty(volume_by_symbol_side),
    }


# -----------------------------
# Excel Export / Styling
# -----------------------------

SHEET_SUMMARY = "Сводка"
SHEET_MATCHES = "Совпадения"
SHEET_MISSING = "Нет в Unity"
SHEET_EXTRA = "Лишнее в Unity"
SHEET_UNITY_STATUS = "Статус Unity"
SHEET_VOL_SYMBOL = "Объем Инстр"
SHEET_VOL_SYMBOL_SIDE = "Объем Инстр+Side"


def _clean_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out = out.replace([np.inf, -np.inf], np.nan)
    for c in out.columns:
        if out[c].dtype == object:
            out[c] = out[c].replace({"nan": "", "NaN": "", "NaT": "", "None": ""})
    return out


def _export_report_xlsx(
    report_path: Path,
    summary: ReconcileSummary,
    params: ReconcileParams,
    exchange_name: str,
    matched_pretty: pd.DataFrame,
    missing_pretty: pd.DataFrame,
    extra_pretty: pd.DataFrame,
    ex_status_pretty: pd.DataFrame,
    unity_status_pretty: pd.DataFrame,
    volume_by_symbol_pretty: Optional[pd.DataFrame],
    volume_by_symbol_side_pretty: Optional[pd.DataFrame],
    top_diffs_strict: Optional[pd.DataFrame] = None,
    top_diffs_notional: Optional[pd.DataFrame] = None,
    raw_exchange: Optional[pd.DataFrame] = None,
    raw_unity: Optional[pd.DataFrame] = None,
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)

    summary_df = pd.DataFrame(
        {
            "Показатель": [
                f"Строк {exchange_name}",
                "Строк Unity",
                "Совпало STRICT",
                "Совпало FUZZY",
                "Совпало NOTIONAL (qty*price)",
                "Нет в Unity (есть в бирже)",
                "Лишнее в Unity (нет в бирже)",
                f"Объем: символов {exchange_name}",
                "Объем: символов Unity",
                "Объем: OK",
                "Объем: Расхождение",
                f"Объем: Только {exchange_name}",
                "Объем: Только Unity",
                f"Объем: Σ Qty {exchange_name}",
                "Объем: Σ Qty Unity",
                f"Объем: Σ Notional {exchange_name}",
                "Объем: Σ Notional Unity",
                f"Диапазон времени {exchange_name} (UTC)",
                "Диапазон времени Unity (UTC)",
                "Warning",
            ],
            "Значение": [
                summary.rows_exchange,
                summary.rows_unity,
                summary.matched_strict,
                summary.matched_fuzzy,
                summary.matched_notional,
                summary.missing_in_unity,
                summary.extra_in_unity,
                summary.volume_symbols_exchange,
                summary.volume_symbols_unity,
                summary.volume_symbols_ok,
                summary.volume_symbols_diff,
                summary.volume_symbols_only_exchange,
                summary.volume_symbols_only_unity,
                summary.volume_total_qty_exchange,
                summary.volume_total_qty_unity,
                summary.volume_total_notional_exchange,
                summary.volume_total_notional_unity,
                summary.exchange_time_range_utc,
                summary.unity_time_range_utc,
                summary.warning,
            ],
        }
    )

    sheet_exchange_status = f"Статус {exchange_name}"[:31]

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        _clean_df_for_excel(summary_df).to_excel(writer, sheet_name=SHEET_SUMMARY, index=False)

        _clean_df_for_excel(matched_pretty).to_excel(writer, sheet_name=SHEET_MATCHES, index=False)
        _clean_df_for_excel(missing_pretty).to_excel(writer, sheet_name=SHEET_MISSING, index=False)
        _clean_df_for_excel(extra_pretty).to_excel(writer, sheet_name=SHEET_EXTRA, index=False)

        _clean_df_for_excel(ex_status_pretty).to_excel(writer, sheet_name=sheet_exchange_status, index=False)
        _clean_df_for_excel(unity_status_pretty).to_excel(writer, sheet_name=SHEET_UNITY_STATUS, index=False)

        if volume_by_symbol_pretty is not None:
            _clean_df_for_excel(volume_by_symbol_pretty).to_excel(writer, sheet_name=SHEET_VOL_SYMBOL, index=False)
        if volume_by_symbol_side_pretty is not None:
            _clean_df_for_excel(volume_by_symbol_side_pretty).to_excel(writer, sheet_name=SHEET_VOL_SYMBOL_SIDE, index=False)

        if params.export_debug_sheets:
            if top_diffs_strict is not None:
                _clean_df_for_excel(top_diffs_strict).to_excel(writer, sheet_name="TopDiffs STRICT"[:31], index=False)
            if top_diffs_notional is not None:
                _clean_df_for_excel(top_diffs_notional).to_excel(writer, sheet_name="TopDiffs Notional"[:31], index=False)
            if raw_exchange is not None:
                _clean_df_for_excel(raw_exchange).to_excel(writer, sheet_name=f"RAW {exchange_name}"[:31], index=False)
            if raw_unity is not None:
                _clean_df_for_excel(raw_unity).to_excel(writer, sheet_name="RAW Unity"[:31], index=False)

    _style_workbook(report_path, params)


def _apply_excel_table(ws) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    safe = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)
    name = ("T_" + safe)[:250]
    tab = Table(displayName=name, ref=ws.dimensions)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _style_workbook(report_path: Path, params: ReconcileParams) -> None:
    wb = load_workbook(report_path)

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue
        ws.freeze_panes = "A2" if ws.max_row >= 2 else "A1"
        _apply_header_style(ws)
        if ws.max_row >= 2 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        _apply_excel_table(ws)
        _apply_column_formats(ws, params)
        _apply_conditional_styles(ws)
        _autosize_columns(ws)

    wb.save(report_path)


def _apply_header_style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=False)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22


def _find_col_by_header(ws, header_name: str) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and v.strip() == header_name:
            return col
    return None


def _apply_column_formats(ws, params: ReconcileParams) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return

    qty_fmt = "0." + ("0" * max(0, int(params.qty_decimals)))
    price_fmt = "0." + ("0" * max(0, int(params.price_decimals)))
    notional_fmt = "0." + ("0" * max(0, int(params.notional_decimals)))
    score_fmt = "0.00"
    dt_fmt = "yyyy-mm-dd hh:mm:ss"

    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if not isinstance(header, str):
            continue
        h = header.strip()

        if ("UTC" in h) or ("Время" in h) or ("Time" in h):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = dt_fmt
            continue

        if h in {"Score", "Δt_sec"}:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = score_fmt
            continue

        if ("Qty" in h) or (h == "ΔQty"):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = qty_fmt
            continue

        if ("Цена" in h) or ("Price" in h):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = price_fmt
            continue

        if ("Объем" in h) or (h == "ΔОбъем") or ("Notional" in h) or ("Volume" in h):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = notional_fmt
            continue


def _apply_conditional_styles(ws) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return

    full_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

    fill_ok = PatternFill("solid", fgColor="E8F5E9")
    fill_warn = PatternFill("solid", fgColor="FFF8E1")
    fill_bad = PatternFill("solid", fgColor="FFEBEE")
    fill_info = PatternFill("solid", fgColor="E3F2FD")

    status_col = _find_col_by_header(ws, "Статус")
    if status_col:
        colL = get_column_letter(status_col)
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colL}2="OK"'], fill=fill_ok))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colL}2="Расхождение"'], fill=fill_bad))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,5)="Только"'], fill=fill_warn))

        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,7)="СОВПАЛО"'], fill=fill_ok))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,4)="НЕТ_"'], fill=fill_bad))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'LEFT(${colL}2,7)="ЛИШНЕЕ_"'], fill=fill_warn))

    mt_col = _find_col_by_header(ws, "Тип_совпадения")
    if mt_col:
        colM = get_column_letter(mt_col)
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colM}2="STRICT"'], fill=fill_ok))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colM}2="FUZZY"'], fill=fill_warn))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[f'${colM}2="NOTIONAL"'], fill=fill_info))


def _autosize_columns(ws, sample_rows: int = 2000) -> None:
    if ws.max_column < 1:
        return
    max_r = min(ws.max_row, sample_rows)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 8
        for row in range(1, max_r + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > best:
                best = len(s)
        ws.column_dimensions[letter].width = min(max(best + 2, 10), 60)


# -----------------------------
# Core reconcile (internal)
# -----------------------------

def _reconcile_core(
    unity_xlsx_path: Path,
    exchange_path: Path,
    exchange_type: str,
    params: ReconcileParams,
) -> Tuple[str, str, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, Optional[Dict[str, float]], int]:
    """
    Returns:
      exchange_name, exchange_code, unity_raw, exchange_raw_std, unity_n, exchange_n, contract_map, used_unity_offset
    """
    exchange_type = (exchange_type or "BINANCE").upper().strip()
    if exchange_type not in {"BINANCE", "OKX", "BYBIT"}:
        raise ValueError(f"Unsupported exchange_type: {exchange_type}")

    unity_raw = pd.read_excel(unity_xlsx_path)
    unity_n, used_unity_offset = _normalize_unity(unity_raw, params)

    if exchange_type == "BINANCE":
        exchange_name = "Binance"
        exchange_code = "B"
        exchange_offset = 0  # Binance exports often UTC
        raw0 = _read_binance_file(exchange_path, params.binance_delimiter)
        exchange_raw = _prepare_binance_to_standard(raw0)
        exchange_n = _normalize_exchange_common(
            exchange_raw,
            params,
            time_offset_hours=exchange_offset,
            symbol_mapper=_extract_symbol_basic,
            action_filter={"BUY", "SELL"},
        )
        return exchange_name, exchange_code, unity_raw, exchange_raw, unity_n, exchange_n, None, used_unity_offset

    if exchange_type == "BYBIT":
        exchange_name = "Bybit"
        exchange_code = "Y"
        exchange_offset = int(params.bybit_utc_offset_hours or 0)
        raw0 = _read_bybit_file(exchange_path)
        exchange_raw = _prepare_bybit_to_standard(raw0)
        exchange_n = _normalize_exchange_common(
            exchange_raw,
            params,
            time_offset_hours=exchange_offset,
            symbol_mapper=_extract_symbol_basic,
            action_filter={"BUY", "SELL"} if params.bybit_filter_trade_actions else None,
        )
        return exchange_name, exchange_code, unity_raw, exchange_raw, unity_n, exchange_n, None, used_unity_offset

    # OKX
    exchange_name = "OKX"
    exchange_code = "O"

    okx_df, tz = _read_okx_xlsx(exchange_path)
    exchange_raw = _prepare_okx_to_standard(okx_df)

    detected = tz
    exchange_offset = int(params.okx_utc_offset_hours) if params.okx_utc_offset_hours is not None else (int(detected) if detected is not None else 0)
    action_filter = {"BUY", "SELL"} if params.okx_filter_trade_actions else None

    contract_map = _infer_okx_contract_value_map(unity_n, exchange_raw, _extract_symbol_from_okx, action_filter, params)
    trading_unit_col = "Trading Unit" if "Trading Unit" in exchange_raw.columns else None

    exchange_n = _normalize_exchange_common(
        exchange_raw,
        params,
        time_offset_hours=exchange_offset,
        symbol_mapper=_extract_symbol_from_okx,
        action_filter=action_filter,
        contract_value_map=contract_map,
        trading_unit_col=trading_unit_col,
    )

    return exchange_name, exchange_code, unity_raw, exchange_raw, unity_n, exchange_n, contract_map, used_unity_offset


# -----------------------------
# Public API
# -----------------------------

def reconcile_to_report(
    unity_xlsx_path: Path,
    exchange_path: Path,
    report_dir: Path,
    exchange_type: str = "BINANCE",
    params: Optional[ReconcileParams] = None,
) -> ReconcileResult:
    params = params or ReconcileParams()

    exchange_name, exchange_code, unity_raw, exchange_raw, unity_n, exchange_n, contract_map, used_unity_offset = _reconcile_core(
        unity_xlsx_path=unity_xlsx_path,
        exchange_path=exchange_path,
        exchange_type=exchange_type,
        params=params,
    )

    # STRICT (key: symbol|side|qty|price) — duplicates matched by time OR by original order
    matched_strict, missing_in_unity, extra_in_unity = _reconcile_multiset_by_key(
        unity_n, exchange_n, "match_key",
        sort_by_time=params.match_duplicates_by_time,
    )

    # FUZZY
    matched_fuzzy = pd.DataFrame(columns=["exchange_idx", "unity_idx", "score"])
    if params.enable_fuzzy:
        matched_fuzzy, missing_in_unity, extra_in_unity = _reconcile_fuzzy(missing_in_unity, extra_in_unity, params)

    # NOTIONAL fallback (key: symbol|side|[minute]|qty*price)
    matched_notional = pd.DataFrame(columns=["exchange_idx", "unity_idx", "key"])
    if params.enable_notional_fallback:
        matched_notional, missing_in_unity, extra_in_unity = _reconcile_multiset_by_key(
            extra_in_unity, missing_in_unity, "notional_key",
            sort_by_time=params.match_duplicates_by_time,
        )

    parts = []
    if not matched_strict.empty:
        parts.append(matched_strict.assign(match_type="STRICT").rename(columns={"key": "key_used"}))
    if not matched_fuzzy.empty:
        parts.append(matched_fuzzy.assign(match_type="FUZZY").assign(key_used=""))
    if not matched_notional.empty:
        parts.append(matched_notional.assign(match_type="NOTIONAL").rename(columns={"key": "key_used"}))

    matched_all = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(
        columns=["exchange_idx", "unity_idx", "match_type", "key_used", "score"]
    )

    ex_status = exchange_n.copy()
    ex_status["status"] = "НЕТ_В_UNITY"
    ex_status["matched_unity_idx"] = np.nan

    uni_status = unity_n.copy()
    uni_status["status"] = "ЛИШНЕЕ_В_UNITY"
    uni_status["matched_exchange_idx"] = np.nan

    def _apply_matches(mdf: pd.DataFrame, label: str) -> None:
        if mdf is None or mdf.empty:
            return
        for _, r in mdf.iterrows():
            eidx = int(r["exchange_idx"])
            uidx = int(r["unity_idx"])
            ex_status.loc[eidx, "status"] = f"СОВПАЛО_{label}"
            ex_status.loc[eidx, "matched_unity_idx"] = uidx
            uni_status.loc[uidx, "status"] = f"СОВПАЛО_{label}"
            uni_status.loc[uidx, "matched_exchange_idx"] = eidx

    _apply_matches(matched_strict, "STRICT")
    _apply_matches(matched_fuzzy, "FUZZY")
    _apply_matches(matched_notional, "ОБЪЕМ")

    ex_range = f"{exchange_n['trade_dt_utc'].min()} → {exchange_n['trade_dt_utc'].max()}"
    u_range = f"{unity_n['trade_dt_utc'].min()} → {unity_n['trade_dt_utc'].max()}"

    warning = ""
    ex_min = exchange_n["trade_dt_utc"].min()
    ex_max = exchange_n["trade_dt_utc"].max()
    u_min = unity_n["trade_dt_utc"].min()
    u_max = unity_n["trade_dt_utc"].max()

    if pd.notna(ex_min) and pd.notna(ex_max) and pd.notna(u_min) and pd.notna(u_max):
        if u_min < ex_min or u_max > ex_max:
            warning = (
                f"Unity time range выходит за диапазон {exchange_name}. "
                f"Проверь период выгрузки/таймзону. "
                f"Unity UTC offset: {used_unity_offset}"
            )

    if contract_map:
        warning = (warning + " | " if warning else "") + f"OKX contracts→base: {contract_map}"

    volume_by_symbol = None
    volume_by_symbol_side = None
    vol_symbols_exchange = 0
    vol_symbols_unity = 0
    vol_ok = 0
    vol_diff = 0
    vol_only_exchange = 0
    vol_only_unity = 0

    if params.enable_volume_recon:
        agg_ex_sym = _agg_volume(exchange_n, by_side=False)
        agg_u_sym = _agg_volume(unity_n, by_side=False)
        volume_by_symbol = _compare_volume(agg_ex_sym, agg_u_sym, by_side=False, params=params)

        vol_symbols_exchange = int(agg_ex_sym["symbol"].nunique())
        vol_symbols_unity = int(agg_u_sym["symbol"].nunique())
        vol_ok = int((volume_by_symbol["status"] == "OK").sum())
        vol_diff = int((volume_by_symbol["status"] == "Расхождение").sum())
        vol_only_exchange = int((volume_by_symbol["status"] == "Только Биржа").sum())
        vol_only_unity = int((volume_by_symbol["status"] == "Только Unity").sum())

        if params.volume_group_by_side:
            agg_ex_ss = _agg_volume(exchange_n, by_side=True)
            agg_u_ss = _agg_volume(unity_n, by_side=True)
            volume_by_symbol_side = _compare_volume(agg_ex_ss, agg_u_ss, by_side=True, params=params)

    vol_total_qty_ex = float(np.nansum(exchange_n["qty"]))
    vol_total_qty_u = float(np.nansum(unity_n["qty"]))
    vol_total_not_ex = float(np.nansum(exchange_n["notional"]))
    vol_total_not_u = float(np.nansum(unity_n["notional"]))

    summary = ReconcileSummary(
        exchange_name=exchange_name,
        rows_exchange=int(len(exchange_n)),
        rows_unity=int(len(unity_n)),

        matched_strict=int(len(matched_strict)),
        matched_fuzzy=int(len(matched_fuzzy)),
        matched_notional=int(len(matched_notional)),

        missing_in_unity=int(len(missing_in_unity)),
        extra_in_unity=int(len(extra_in_unity)),

        volume_symbols_exchange=vol_symbols_exchange,
        volume_symbols_unity=vol_symbols_unity,
        volume_symbols_ok=vol_ok,
        volume_symbols_diff=vol_diff,
        volume_symbols_only_exchange=vol_only_exchange,
        volume_symbols_only_unity=vol_only_unity,
        volume_total_qty_exchange=_qround_float(vol_total_qty_ex, params.qty_decimals),
        volume_total_qty_unity=_qround_float(vol_total_qty_u, params.qty_decimals),
        volume_total_notional_exchange=_qround_float(vol_total_not_ex, params.notional_decimals),
        volume_total_notional_unity=_qround_float(vol_total_not_u, params.notional_decimals),

        exchange_time_range_utc=ex_range,
        unity_time_range_utc=u_range,
        warning=warning,
    )

    pretty = _build_pretty_tables(
        matched_all=matched_all,
        exchange_n=exchange_n,
        unity_n=unity_n,
        missing_in_unity=missing_in_unity,
        extra_in_unity=extra_in_unity,
        ex_status=ex_status,
        uni_status=uni_status,
        volume_by_symbol=volume_by_symbol,
        volume_by_symbol_side=volume_by_symbol_side,
        exchange_name=exchange_name,
        exchange_code=exchange_code,
        params=params,
    )

    report_id = str(uuid.uuid4())
    report_path = report_dir / f"unity_vs_{exchange_name.lower()}_{report_id}.xlsx"

    top_diffs_strict = _top_key_diffs(unity_n, exchange_n, "match_key", limit=50) if params.export_debug_sheets else None
    top_diffs_notional = _top_key_diffs(unity_n, exchange_n, "notional_key", limit=50) if params.export_debug_sheets else None

    _export_report_xlsx(
        report_path=report_path,
        summary=summary,
        params=params,
        exchange_name=exchange_name,
        matched_pretty=pretty["matched"],          # type: ignore[arg-type]
        missing_pretty=pretty["missing"],          # type: ignore[arg-type]
        extra_pretty=pretty["extra"],              # type: ignore[arg-type]
        ex_status_pretty=pretty["ex_status"],      # type: ignore[arg-type]
        unity_status_pretty=pretty["uni_status"],  # type: ignore[arg-type]
        volume_by_symbol_pretty=pretty["vol_sym"],
        volume_by_symbol_side_pretty=pretty["vol_ss"],
        top_diffs_strict=top_diffs_strict,
        top_diffs_notional=top_diffs_notional,
        raw_exchange=exchange_raw if params.export_debug_sheets else None,
        raw_unity=unity_raw if params.export_debug_sheets else None,
    )

    return ReconcileResult(report_id=report_id, report_path=report_path, summary=summary)


def reconcile_to_report_with_preview(
    unity_xlsx_path: Path,
    exchange_path: Path,
    report_dir: Path,
    exchange_type: str = "BINANCE",
    params: Optional[ReconcileParams] = None,
    preview_limit: int = 2000,
) -> Tuple[ReconcileResult, Dict[str, List[Dict[str, Any]]]]:
    params = params or ReconcileParams()

    res = reconcile_to_report(
        unity_xlsx_path=unity_xlsx_path,
        exchange_path=exchange_path,
        report_dir=report_dir,
        exchange_type=exchange_type,
        params=params,
    )

    preview: Dict[str, List[Dict[str, Any]]] = {
        "matches": [],
        "missing": [],
        "extra": [],
        "exchange_status": [],
        "unity_status": [],
        "volume_symbol": [],
        "volume_symbol_side": [],
    }
    return res, preview