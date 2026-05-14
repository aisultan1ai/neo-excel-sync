import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def build_funding_export(account: dict, records: list, start_date: Optional[str], end_date: Optional[str], symbol: Optional[str]) -> io.BytesIO:
    records_sorted = sorted(
        records, key=lambda r: (str(r.get("date_local", "")), str(r.get("symbol", "")))
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Account"
    ws["B1"] = account["name"]
    ws["A2"] = "Exported"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    if start_date:
        ws["A3"], ws["B3"] = "From", start_date
    if end_date:
        ws["A4"], ws["B4"] = "To", end_date
    if symbol:
        ws["A5"], ws["B5"] = "Symbol", symbol
    ws.append([])

    by_symbol: dict = {}
    for r in records_sorted:
        key = (r["symbol"], r["asset"])
        if key not in by_symbol:
            by_symbol[key] = {"total": 0.0, "count": 0}
        by_symbol[key]["total"] += float(r["income"])
        by_symbol[key]["count"] += 1

    summary_headers = ["Symbol", "Asset", "Total Income", "Records"]
    ws.append(summary_headers)
    for col in range(1, 5):
        c = ws.cell(row=ws.max_row, column=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = border

    for (sym, ast), data in sorted(by_symbol.items()):
        ws.append([sym, ast, round(data["total"], 8), data["count"]])
        for col in range(1, 5):
            ws.cell(row=ws.max_row, column=col).border = border

    grand_total = sum(float(r["income"]) for r in records_sorted)
    ws.append(["TOTAL", "", round(grand_total, 8), len(records_sorted)])
    for col in range(1, 5):
        c = ws.cell(row=ws.max_row, column=col)
        c.font = Font(bold=True)
        c.border = border

    for col, w in zip("ABCD", [20, 10, 18, 12]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Raw Data")
    raw_headers = ["Date", "Symbol", "Asset", "Income", "UTC Time"]
    ws2.append(raw_headers)
    for col in range(1, 6):
        c = ws2.cell(row=1, column=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = border

    for r in records_sorted:
        dt_utc = r.get("datetime_utc")
        dt_str = dt_utc.strftime("%Y-%m-%d %H:%M:%S") if dt_utc else ""
        ws2.append([str(r.get("date_local", "")), r["symbol"], r["asset"], round(float(r["income"]), 8), dt_str])
        for col in range(1, 6):
            ws2.cell(row=ws2.max_row, column=col).border = border

    for col, w in zip("ABCDE", [12, 20, 8, 16, 20]):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_export_filename(account: dict, start_date: Optional[str], end_date: Optional[str]) -> str:
    parts = [f"funding_{account['name']}"]
    if start_date:
        parts.append(start_date)
    if end_date:
        parts.append(end_date)
    return "_".join(parts) + ".xlsx"
